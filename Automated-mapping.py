import logging
import advertools as adv
import pandas as pd
import streamlit as st
import time
from openpyxl import load_workbook
from polyfuzz import PolyFuzz
from polyfuzz.models import RapidFuzz

logging.basicConfig(format="%(asctime)s — %(levelname)s — %(name)s — %(message)s", level=logging.INFO)

matcher = RapidFuzz(n_jobs=1, score_cutoff=0.80)
model = PolyFuzz(matcher)

hide_devcontainer_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            </style>
            """

st.markdown("""
<style>
.big-font {
    font-size:50px !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<h1>Mappingator V1.0</h1>
<p>Avec cet outil, vous serez en mesure de grandement accélérer le jumelage des URLs d'un site actuel vers sa nouvelle version.</p>
<h4>Instructions:</h4>
<ul>
<li>À l'aide de Screaming Frog, bâtissez-vous deux liste d'URLs. Une avec les anciennes URLs et l'autre avec les nouvelles URLs. Le titre de ces documents n'a aucune importance</li>
<li>Téléversez le fichier "Legacy Crawl" des URLs (xlsx)</li>
<li>Téléversez le fichier "New Crawl" des URLs (xlsx)</li>
<li>Téléchargez le fichier (xlsx) après que Chuck ait fini de démonter ses ennemis. Si ses ennemis sont nombreux, ça risque de prendre quelques minutes.</li>
</ul>
<h4>Chuck ne va pas être capable de vaincre ses ennemis si les consignes du bas ne sont pas respectées: </h4>
<ul>
<li>La colonne 1 doit se nommer "Address" et contenir les URLs complètes, incluant le "http(s)://"</li>
<li>En plus de la colonne "Address" Chuck à besoin de colonnes supplémentaires. Les titres des colonnes supplémentaires doivent exister dans les deux fichiers, même si les cellules des colonnes sont vides:
 <ul>
 <li>"Title 1" "H1-1" "H2-1"</li>
 </ul>

</ul>
<br>
""", unsafe_allow_html=True)

legacy_file = st.file_uploader('Téléverser le crawl des LEGACY URLs', type='xlsx', key='legacy')

input_files = []
crawl_columns = ['Address','Title 1','H1-1','H2-1']


def analyze_crawls(crawls):
    with st.spinner('Processing site crawls. . .'):
        progress_bar = st.progress(0)
        for crawl_index, crawl in enumerate(crawls):
            wb = load_workbook(filename=crawl)
            sheet_name = wb.sheetnames
            input_files.append([crawl, sheet_name])
            progress_bar.progress((crawl_index + 1) / len(crawls))
            time.sleep(0.01)

        legacy_crawl = pd.read_excel(input_files[0][0], sheet_name=input_files[0][1][0])
        legacy_crawl = legacy_crawl[crawl_columns]
        new_crawl = pd.read_excel(input_files[1][0], sheet_name=input_files[1][1][0])
        new_crawl = new_crawl[crawl_columns]
        legacy_col_1 = list(legacy_crawl.columns)[0]
        new_col_1 = list(new_crawl.columns)[0]
        legacy_urls = legacy_crawl[legacy_col_1].tolist()
        new_urls = new_crawl[new_col_1].tolist()
    url_parse(legacy_urls, legacy_crawl, new_urls, new_crawl)


def url_match(legacy_paths, new_paths, legacy_url_parse, new_url_parse):
    with st.spinner('Analyzing URL Paths. . .'):
        model.match(legacy_paths, new_paths)
        pfuzz_df = model.get_matches()
        pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
        pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
        pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .800]
        print(pfuzz_df.head())

        join_df = pd.merge(pfuzz_df, legacy_url_parse, left_on='From', right_on='path')
        join_df_2 = pd.merge(join_df, new_url_parse, left_on='To', right_on='path')
        join_df_2.rename(
            columns={'url_x': 'Legacy URL', 'url_y': 'New URL', 'path_x': 'Legacy URL Path', 'path_y': 'New URL Path'},
            inplace=True)
        url_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL Path', 'New URL Path', 'Legacy URL', 'New URL']]
        url_df = url_df.drop_duplicates()
        url_df.head()
    return url_df


def slug_match(legacy_slugs, new_slugs, legacy_url_parse, new_url_parse):
    with st.spinner('Analyzing URL Slugs. . .'):
        model.match(legacy_slugs, new_slugs)

        pfuzz_df = model.get_matches()
        pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
        pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
        pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .800]
        print(pfuzz_df.head())

        join_df = pd.merge(pfuzz_df, legacy_url_parse, left_on='From', right_on='last_dir')
        join_df_2 = pd.merge(join_df, new_url_parse, left_on='To', right_on='last_dir')
        join_df_2.rename(
            columns={'url_x': 'Legacy URL', 'url_y': 'New URL', 'path_x': 'Legacy URL Path', 'path_y': 'New URL Path'},
            inplace=True)
        slug_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL Path', 'New URL Path', 'Legacy URL', 'New URL']]
        slug_df = slug_df.drop_duplicates()
        slug_df.head()
    return slug_df


def title_match(legacy_titles, new_titles, legacy_crawl, new_crawl):
    with st.spinner('Analyzing Title tags. . .'):
        model.match(legacy_titles, new_titles)

        pfuzz_df = model.get_matches()
        pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
        pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
        pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .700]
        print(pfuzz_df.head())

        join_df = pd.merge(pfuzz_df, legacy_crawl, left_on='From', right_on='Title 1')
        join_df_2 = pd.merge(join_df, new_crawl, left_on='To', right_on='Title 1').drop_duplicates()
        join_df_2.rename(columns={'Address_x': 'Legacy URL', 'Address_y': 'New URL'}, inplace=True)
        title_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL', 'New URL']]
        title_df = title_df.drop_duplicates()
        print(title_df.head())
    return title_df


def h1_match(legacy_h1, new_h1, legacy_crawl, new_crawl):
    with st.spinner('Analyzing h1 tags. . .'):
        model.match(legacy_h1, new_h1)

        pfuzz_df = model.get_matches()
        pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
        pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
        pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .900]
        print(pfuzz_df.head())

        join_df = pd.merge(pfuzz_df, legacy_crawl, left_on='From', right_on='H1-1')
        join_df_2 = pd.merge(join_df, new_crawl, left_on='To', right_on='H1-1')
        join_df_2.rename(columns={'Address_x': 'Legacy URL', 'Address_y': 'New URL'}, inplace=True)
        h1_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL', 'New URL']]
        h1_df = h1_df.drop_duplicates()
    return h1_df


def h2_match(legacy_h2, new_h2, legacy_crawl, new_crawl):
    with st.spinner('Analyzing h2 tags. . .'):
        model.match(legacy_h2, new_h2)

        pfuzz_df = model.get_matches()
        pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
        pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
        pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .900]
        print(pfuzz_df.head())

        join_df = pd.merge(pfuzz_df, legacy_crawl, left_on='From', right_on='H2-1')
        join_df_2 = pd.merge(join_df, new_crawl, left_on='To', right_on='H2-1').drop_duplicates()
        join_df_2.rename(columns={'Address_x': 'Legacy URL', 'Address_y': 'New URL'}, inplace=True)
        h2_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL', 'New URL']]
        h2_df = h2_df.drop_duplicates()
        print(h2_df.head())
    return h2_df


def url_parse(legacy_urls, legacy_crawl, new_urls, new_crawl):
    with st.spinner('Deconstructing URLs. . .'):
        url_parse_cols = ['url', 'path', 'last_dir']
        legacy_url_parse = adv.url_to_df(legacy_urls)
        legacy_url_parse = legacy_url_parse[url_parse_cols]
        new_url_parse = adv.url_to_df(new_urls)
        new_url_parse = new_url_parse[url_parse_cols]

        legacy_paths = legacy_url_parse['path']
        new_paths = new_url_parse['path']
        legacy_slug = legacy_url_parse['last_dir']
        new_slug = new_url_parse['last_dir']
        legacy_titles = legacy_crawl['Title 1']
        new_titles = new_crawl['Title 1']
        legacy_h1 = legacy_crawl['H1-1']
        new_h1 = new_crawl['H1-1']
        legacy_h2 = legacy_crawl['H2-1']
        new_h2 = new_crawl['H2-1']
    match_dfs = [
        url_match(legacy_paths, new_paths, legacy_url_parse, new_url_parse),
        slug_match(legacy_slug, new_slug, legacy_url_parse, new_url_parse),
        title_match(legacy_titles, new_titles, legacy_crawl, new_crawl),
        h1_match(legacy_h1, new_h1, legacy_crawl, new_crawl),
        h2_match(legacy_h2, new_h2, legacy_crawl, new_crawl)
    ]
    export_dfs(match_dfs)


def export_dfs(match_dfs):
    sheet_names = ['URL Match', 'Slug Match', 'Title Match', 'H1 Match', 'H2 Match']
    with pd.ExcelWriter('mapped_urls.xlsx') as writer:
        for df in enumerate(match_dfs):
            print(df[1])
            df[1].to_excel(writer, sheet_name=sheet_names[df[0]], index=False)

    my_file = pd.read_excel('mapped_urls.xlsx')

    with open("mapped_urls.xlsx", "rb") as file:
        st.download_button(label='Download Match Analysis',
                           data=file,
                           file_name='mapped_urls.xlsx',
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    if legacy_file is not None:
        new_file = st.file_uploader('Téléversez le crawl des NEW URLs', type='xlsx', key='new')
        if new_file is not None:
            crawl_files = [legacy_file, new_file]
            analyze_crawls(crawl_files)

st.write('---')
st.markdown(hide_devcontainer_style, unsafe_allow_html=True)
st.write('Auteur: [Rablab](https://rablab.ca) | Agence de performance SEO | [Contactez-nous!](https://rablab.ca/contact/)️⚡')
